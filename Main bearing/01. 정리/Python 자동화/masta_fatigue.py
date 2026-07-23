"""
MASTA 메인베어링 피로 시계열 반복해석 + 손상 계산 (방식 A)
========================================================
GH Bladed 시계열(.$150) 각 시간지점을 1점씩 로드케이스에 입력하고
System Deflection 해석 → 베어링 수명/응력 추출 → 점별 손상 계산 → CSV 누적.
Miner 손상은 회전수·발생빈도(FatigueHours.txt)로 외부 계산(요약 시트 별도).
(메모리 안전: 모델 1회 로드, 값만 갱신, 결과는 1행씩 즉시 기록)

좌표 변환 (파일 hub 6분력 → MASTA 입력):
    Point Load 5성분: force_x=-Fz*1000, force_y=+Fy*1000, axial_load=+Fx*1000 (N)
                       moment_x=-Mz*1000, moment_y=+My*1000 (Nm)
    축토크(6번째): Input Power Load torque = +Mx*1000 (Nm, 파일 Mx=MASTA Mz) → 샤프트 비틀림
속도: 파일 rpm → Input Power Load.speed (rad/s)
샤프트: DIN743:2012-12 무한수명 피로안전율 점별 저장, 최소값·임계시점 요약
단위: MASTA API=SI(N, Nm, rad/s). 응력 Pa, 수명 사이클(Rev).

실행: PYTHONUTF8=1 python -X utf8 masta_fatigue.py
"""
import csv
import math
import os
import traceback

# ── 설정: fatigue_config.xlsx(시트 '설정')에서 읽음. 없으면 아래 기본값 ──
HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_XLSX = os.path.join(HERE, "fatigue_config.xlsx")


def load_config(path):
    """설정 xlsx(시트 '설정': A=파라미터, B=값) → dict. 없거나 실패 시 {}."""
    cfg = {}
    if not os.path.exists(path):
        print(f"[설정] {os.path.basename(path)} 없음 → 기본값 사용")
        return cfg
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb["설정"] if "설정" in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            k = str(row[0]).strip()
            v = row[1] if len(row) > 1 else None
            if k and not k.startswith("#"):
                cfg[k] = v
        print(f"[설정] {os.path.basename(path)} 로드 ({len(cfg)}개 항목)")
    except Exception as e:
        print("[설정] 읽기 실패 → 기본값 사용:", e)
    return cfg


CFG = load_config(CONFIG_XLSX)


def cfg_get(key, default):
    v = CFG.get(key)
    if v is None or (isinstance(v, str) and not v.strip()):
        return default
    return v


MASTA_DIR   = r"C:\Program Files\SMT\MASTA 14.1.1"
MODEL_PATH  = str(cfg_get("MODEL_PATH", r"D:\AI\AI_Seminar\Main bearing\02. 자료\MASTA\26MW_메인베어링_기본설계_v1.3_샤프트 두께,형상 2안_피로하중 반영_260720.Masta"))
DLC_FILE    = str(cfg_get("DLC_FILE", r"D:\AI\AI_Seminar\Main bearing\02. 자료\260714 유니슨 피로하중\DLC1.2-c-s1.$150"))
FATIGUE_HRS = str(cfg_get("FATIGUE_HRS", r"D:\AI\AI_Seminar\Main bearing\02. 자료\260714 유니슨 피로하중\FatigueHours.txt"))
DRIVER_LC   = str(cfg_get("DRIVER_LC", "Load Case 1"))
DT_S        = float(cfg_get("DT_S", 0.1))     # 해석 dt(초): 0.1=전량6001, 0.4=1501, 60=11 …
OUT_CSV     = cfg_get("OUT_CSV", None)        # 파일명(비우면 자동명명). 전체경로면 OUT_DIR보다 우선
OUT_DIR     = cfg_get("OUT_DIR", None)        # 저장 폴더(비우면 스크립트 폴더)
MAKE_XLSX   = str(cfg_get("MAKE_XLSX", "Y")).strip().upper().startswith("Y")
DESIGN_YEARS = 30.0
SPEED_AS_RAD_PER_S = True
DT0 = 0.1                                     # 원본 시계열 간격(초)
# ─────────────────────────────────────────────────────────

import masta_clr_legacy  # noqa: F401
import mastapy
mastapy.init(MASTA_DIR)
from mastapy.system_model import Design
from mastapy.system_model.analyses_and_results.static_loads import AnalysisType

RPM2RADS = 2.0 * math.pi / 60.0

# 손상 4종: (컬럼명, L10 사이클 경로, 표준, 수명기준)
DAMAGE_DEFS = [
    ("dmg_ISO281_basic",   "iso2812007.basic_rating_life_cycles",              "ISO281",   "basic"),
    ("dmg_ISO281_mod",     "iso2812007.modified_rating_life_cycles",           "ISO281",   "modified"),
    ("dmg_ISO16281_basic", "isots162812008.basic_reference_rating_life_cycles","ISO16281", "basic"),
    ("dmg_ISO16281_mod",   "isots162812008.modified_reference_rating_life_cycles","ISO16281","modified"),
]
DMG_COLS = [d[0] for d in DAMAGE_DEFS]
# 6분력: 앞 5개=Point Load, Moment_z_Nm=축토크(Input Power Load 토크로 인가)
LOAD_COLS = ["force_x_N", "force_y_N", "axial_load_N", "moment_x_Nm", "moment_y_Nm", "Moment_z_Nm"]
SHAFT_COLS = ["shaft_DIN743_SF_inf", "shaft_SF_offset_m"]


def parse_dlc(path):
    rows = []
    with open(path, "r", encoding="latin-1") as f:
        for ln in f.readlines()[4:]:      # 헤더 4줄 이후
            p = ln.split()
            if len(p) < 8:
                continue
            try:
                v = [float(x) for x in p[:8]]
            except ValueError:
                continue
            rows.append(dict(t=v[0], rpm=v[1], Mx=v[2], My=v[3], Mz=v[4],
                             Fx=v[5], Fy=v[6], Fz=v[7]))
    return rows


def read_scale_factor(path, dlc_name):
    """FatigueHours.txt 에서 DLC의 Scale Factor와 30년 시간(h). 숫자 '4.504.E+04' 보정."""
    try:
        with open(path, encoding="latin-1") as f:
            for ln in f.readlines()[1:]:
                c = ln.rstrip("\n").split("\t")
                if len(c) >= 3 and c[0].strip() == dlc_name:
                    sf = float(c[1].strip().replace(".E", "E"))
                    h30 = float(c[2].strip())
                    return sf, h30
    except Exception as e:
        print("  [warn] FatigueHours 읽기 실패:", e)
    return None, None


def g(obj, path, default=None):
    cur = obj
    for part in path.split("."):
        try:
            cur = getattr(cur, part)
        except Exception:
            return default
        if cur is None:
            return default
    return cur


def fnum(v):
    """float이면 반환, 아니면 None."""
    return v if isinstance(v, (int, float)) else None


def num(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if math.isinf(v):
            return "inf"
        if math.isnan(v):
            return "nan"
    return v


def bname(b):
    n = getattr(b, "name", None)
    return n if isinstance(n, str) else str(b)


def damage(rev, life_cycles):
    """점별 손상 = 회전수 / L10(cycles). 유효하지 않으면 None."""
    lc = fnum(life_cycles)
    if lc is not None and lc > 0 and math.isfinite(lc):
        return rev / lc
    return None


def set_loads(lc, point_load, ipl, rec):
    loads = {
        "force_x_N":    -rec["Fz"] * 1000.0,
        "force_y_N":     rec["Fy"] * 1000.0,
        "axial_load_N":  rec["Fx"] * 1000.0,
        "moment_x_Nm":  -rec["Mz"] * 1000.0,
        "moment_y_Nm":   rec["My"] * 1000.0,
        "Moment_z_Nm":   rec["Mx"] * 1000.0,   # 축토크 = +Mx(=MASTA Mz) → Power Load 토크
    }
    pli = lc.inputs_for_point_load(point_load)
    pli.force_x.force    = loads["force_x_N"]
    pli.force_y.force    = loads["force_y_N"]
    pli.axial_load.force = loads["axial_load_N"]
    pli.moment_x.moment  = loads["moment_x_Nm"]
    pli.moment_y.moment  = loads["moment_y_Nm"]
    pll = lc.inputs_for_power_load(ipl)
    spd = rec["rpm"] * (RPM2RADS if SPEED_AS_RAD_PER_S else 1.0)
    try:
        pll.speed = spd
    except Exception as e:
        print("    [warn] speed 세팅 실패:", str(e).splitlines()[0])
    try:
        pll.torque = loads["Moment_z_Nm"]     # 축토크 인가 → 샤프트 비틀림
    except Exception as e:
        print("    [warn] torque 세팅 실패:", str(e).splitlines()[0])
    return loads


def sample_indices(n_full, dt_s, dt0=DT0):
    """dt(초) → 근사균일 샘플 인덱스(양 끝 포함). 부록 E: N=T/dt+1, idx=round(j·last/(N-1))."""
    last = n_full - 1                     # 6000
    T = last * dt0                        # 600.0 s
    N = int(round(T / float(dt_s))) + 1
    N = max(2, min(N, n_full))
    return [int(round(j * last / (N - 1))) for j in range(N)], N


def quad_weights(idxs, dt0=DT0):
    """사다리꼴 구적 가중 dt[s] (양 끝 절반) → Σw = 전체 구간 T. 비정수 stride도 정확."""
    n = len(idxs)
    if n == 1:
        return [dt0]
    gaps = [(idxs[i + 1] - idxs[i]) * dt0 for i in range(n - 1)]
    w = [0.0] * n
    w[0] = gaps[0] / 2.0
    w[-1] = gaps[-1] / 2.0
    for i in range(1, n - 1):
        w[i] = (gaps[i - 1] + gaps[i]) / 2.0
    return w


def shaft_din743_sf(sd, shaft):
    """샤프트 DIN743:2012-12 무한수명 피로안전율(최소) + 위치 offset(m)."""
    try:
        sres = sd.results_for(shaft)
    except Exception:
        return None, None
    sf = None
    try:
        for it in sres.safety_factors.items:
            try:
                desc = it.description or ""
            except Exception:
                desc = ""
            if "Fatigue Safety Factor for Infinite" in desc:
                sf = fnum(it.safety_factor)
    except Exception:
        pass
    off = fnum(g(sres, "shaft_section_end_with_worst_fatigue_safety_factor_for_infinite_life.offset"))
    return sf, off


# 데이터 시트 헤더 (요청 순서)
DATA_HEADER = (["index", "t_s", "rpm", "rev"] + LOAD_COLS + ["bearing",
               "stress_inner_MPa", "stress_outer_MPa", "s0_static",
               "L10_basic_rev", "L10m_mod_rev", "L10r_rev", "L10mr_rev"]
               + DMG_COLS + SHAFT_COLS)


def write_summary(data_csv, summary_csv, scale_factor, h30, n_points):
    """데이터 CSV 되읽어 (a)베어링 손상 요약 + (b)샤프트 임계점 작성."""
    sums = {}
    crit = None            # 샤프트 DIN743 SF 최소인 행
    min_sf = None
    with open(data_csv, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            b = r["bearing"]
            d = sums.setdefault(b, {k: 0.0 for k in DMG_COLS})
            for k in DMG_COLS:
                try:
                    v = float(r[k])
                    if math.isfinite(v):
                        d[k] += v
                except Exception:
                    pass
            try:
                ssf = float(r["shaft_DIN743_SF_inf"])
                if math.isfinite(ssf) and (min_sf is None or ssf < min_sf):
                    min_sf, crit = ssf, r
            except Exception:
                pass
    with open(summary_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        # (a) 베어링 손상
        w.writerow(["=== (a) 베어링 손상 ==="])
        w.writerow(["bearing", "standard", "life_basis", "sample_damage",
                    "scale_factor", "damage_30yr", "SF_fatigue",
                    "equiv_life_years", "note"])
        for b, dd in sums.items():
            for col, _, std, basis in DAMAGE_DEFS:
                ds = dd[col]
                if scale_factor:
                    d30 = ds * scale_factor
                    sf = (1.0 / d30) if d30 > 0 else ""
                    eq = (DESIGN_YEARS / d30) if d30 > 0 else ""
                else:
                    d30 = sf = eq = ""
                w.writerow([b, std, basis, ds, scale_factor if scale_factor else "",
                            d30, sf, eq, f"{n_points}점 표본(1 DLC)"])
        # (b) 샤프트 임계점 (DIN743 피로 SF 최소)
        w.writerow([])
        w.writerow(["=== (b) 샤프트 임계점 (DIN743 무한수명 피로 SF 최소) ==="])
        cols = (["min_shaft_DIN743_SF_inf", "crit_index", "crit_t_s", "crit_rpm"]
                + ["crit_" + c for c in LOAD_COLS] + ["crit_shaft_SF_offset_m"])
        w.writerow(cols)
        if crit is not None:
            vals = ([min_sf, crit["index"], crit["t_s"], crit["rpm"]]
                    + [crit.get(c, "") for c in LOAD_COLS]
                    + [crit.get("shaft_SF_offset_m", "")])
        else:
            vals = ["" for _ in cols]
        w.writerow(vals)
    print(f"요약 저장: {summary_csv}  (min shaft SF={min_sf})")


def main():
    print("[주의] 대상 .Masta 는 MASTA GUI에서, 출력 CSV/xlsx 는 엑셀에서 닫아주세요.")
    dlc_name = os.path.splitext(os.path.basename(DLC_FILE))[0]
    dt_tag = ("%g" % DT_S).replace(".", "p")
    # 저장 폴더(OUT_DIR) + 파일명(OUT_CSV). OUT_CSV가 전체경로면 그것을 우선.
    out_dir = str(OUT_DIR).strip() if OUT_DIR else HERE
    if not os.path.isabs(out_dir):
        out_dir = os.path.join(HERE, out_dir)
    name = str(OUT_CSV).strip() if OUT_CSV else f"fatigue_{dlc_name.replace('$','')}_dt{dt_tag}.csv"
    if not name.lower().endswith(".csv"):          # 확장자 누락 보정(이름에 점이 있어도 안전)
        name += ".csv"
        print(f"[보정] OUT_CSV 확장자 없음 → '{name}' 사용")
    out_csv = name if os.path.isabs(name) else os.path.join(out_dir, name)
    os.makedirs(os.path.dirname(out_csv), exist_ok=True)
    summary_csv = os.path.splitext(out_csv)[0] + "_summary.csv"
    print(f"저장 폴더: {os.path.dirname(out_csv)}")

    # ── 입력 경로 보정·검증 (설정 엑셀 수기입력 실수 방지) ──
    model_path = MODEL_PATH
    if not os.path.exists(model_path) and os.path.exists(model_path + ".Masta"):
        model_path += ".Masta"
        print("[보정] MODEL_PATH 확장자(.Masta) 자동 추가")
    bad = False
    for label, p in (("MODEL_PATH", model_path), ("DLC_FILE", DLC_FILE),
                     ("FATIGUE_HRS", FATIGUE_HRS)):
        if not os.path.exists(p):
            bad = True
            print(f"[오류] {label} 파일을 찾을 수 없습니다:\n        {p}")
            d = os.path.dirname(p)
            if os.path.isdir(d):
                ext = ".masta" if label == "MODEL_PATH" else os.path.splitext(p)[1].lower()
                cand = [f for f in sorted(os.listdir(d)) if f.lower().endswith(ext)]
                if cand:
                    print(f"        └ 같은 폴더 후보({len(cand)}):")
                    for f in cand[:10]:
                        print("           -", f)
            else:
                print("        └ 폴더 자체가 없습니다. 경로를 확인하세요.")
    if bad:
        print("\n설정 파일(fatigue_config.xlsx)의 경로를 수정한 뒤 다시 실행하세요.")
        raise SystemExit(1)

    print("모델 로드:", model_path)
    design = Design.load(model_path)
    assembly = design.all_parts_of_type_root_assembly()[0]
    point_load = list(assembly.all_parts_of_type_point_load())[0]
    bearings = list(assembly.all_parts_of_type_bearing())
    shaft = list(assembly.all_parts_of_type_shaft())[0]
    ipl = next(p for p in assembly.all_parts_of_type_power_load() if "input" in str(p).lower())
    lc = next(c for c in assembly.design_properties.static_loads
              if getattr(c, "name", "") == DRIVER_LC)

    sf, h30 = read_scale_factor(FATIGUE_HRS, dlc_name)
    print(f"드라이버 LC: {getattr(lc,'name','?')} | 베어링: {[bname(b) for b in bearings]}")
    print(f"DLC={dlc_name} | ScaleFactor={sf} | 30yr Hours={h30}")

    data = parse_dlc(DLC_FILE)
    dt0 = round(data[1]["t"] - data[0]["t"], 6) if len(data) > 1 else DT0
    idxs, n_sel = sample_indices(len(data), DT_S, dt0)
    wts = quad_weights(idxs, dt0)
    print(f"시계열 {len(data)}점(원본 Δt={dt0}s) → 샘플 dt={DT_S}s, "
          f"N={n_sel}점, Σw={sum(wts):.1f}s")
    w_of = dict(zip(idxs, wts))

    done = set()
    if os.path.exists(out_csv):
        with open(out_csv, encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                try:
                    done.add(int(r["index"]))
                except Exception:
                    pass
        print(f"체크포인트: 기존 {len(done)}점 → 스킵")

    new = not os.path.exists(out_csv)
    fcsv = open(out_csv, "a", newline="", encoding="utf-8-sig")
    w = csv.writer(fcsv)
    if new:
        w.writerow(DATA_HEADER)

    n_ok = 0
    for i in idxs:
        if i in done:
            continue
        rec = data[i]
        rev = (rec["rpm"] / 60.0) * w_of[i]     # 점당 회전수 [rev] (사다리꼴 가중 dt)
        try:
            loads = set_loads(lc, point_load, ipl, rec)
            sd = lc.analysis_of(AnalysisType.SYSTEM_DEFLECTION)
            sd.perform_analysis()
            load_vals = [loads[c] for c in LOAD_COLS]
            sh_sf, sh_off = shaft_din743_sf(sd, shaft)      # 샤프트 DIN743 (점당 1값)
            shaft_vals = [num(sh_sf), num(sh_off)]
            for b in bearings:
                res = sd.results_for(b)
                d = g(res, "component_detailed_analysis")
                sin = fnum(g(d, "maximum_normal_stress_inner"))
                sout = fnum(g(d, "maximum_normal_stress_outer"))
                s0 = g(d, "iso762006.safety_factor")
                L10b = g(d, "iso2812007.basic_rating_life_cycles")
                L10m = g(d, "iso2812007.modified_rating_life_cycles")
                L10r = g(d, "isots162812008.basic_reference_rating_life_cycles")
                L10mr = g(d, "isots162812008.modified_reference_rating_life_cycles")
                dmgs = [damage(rev, g(d, path)) for _, path, _, _ in DAMAGE_DEFS]
                row = ([i, rec["t"], rec["rpm"], rev] + load_vals + [bname(b),
                        num(sin / 1e6 if sin is not None else None),
                        num(sout / 1e6 if sout is not None else None),
                        num(s0), num(L10b), num(L10m), num(L10r), num(L10mr)]
                       + [num(x) for x in dmgs] + shaft_vals)
                w.writerow(row)
            fcsv.flush()
            n_ok += 1
            d0 = g(sd.results_for(bearings[0]), "component_detailed_analysis")
            print(f"  [{i:5d}] t={rec['t']:.2f} rpm={rec['rpm']:.3f} rev={rev:.5f} "
                  f"| σin={num(fnum(g(d0,'maximum_normal_stress_inner'))/1e6):.0f}MPa "
                  f"s0={num(g(d0,'iso762006.safety_factor')):.2f} "
                  f"shaftSF={num(sh_sf)}")
        except Exception:
            print(f"  [{i}] 해석 실패:")
            traceback.print_exc()

    fcsv.close()
    print(f"\n데이터 저장: {out_csv}  (신규 {n_ok}점 × {len(bearings)}베어링)")

    # 요약 시트 (데이터 CSV 되읽어 집계 → 재시작에도 안전)
    write_summary(out_csv, summary_csv, sf, h30, n_sel)

    # 엑셀(2시트) 자동 생성
    if MAKE_XLSX:
        try:
            from make_xlsx import build_xlsx
            build_xlsx(out_csv)
        except Exception as e:
            print("  [warn] 엑셀 생성 실패:", e)


if __name__ == "__main__":
    main()
