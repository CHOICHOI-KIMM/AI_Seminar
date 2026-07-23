"""
부록 8 총괄 손상정리 엑셀 — T50 요약 CSV 23개(빈12+점11)를 한 파일로 통합
========================================================================
시트1 "총괄"      : 방식·dt·베어링·표준·기준별 손상 184행 + 편향ε (첨부 양식 + 식별열)
시트2 "주지표요약": 부록 8-5 표 재현 (ISO16281 modified 등가수명 + 편향)
데이터: Results/T50_*_summary.csv  (MASTA 불필요, 재실행 시 최신 반영)
"""
import csv
import glob
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join("Results", "부록8_총괄_손상정리.xlsx")
DLC = "DLC1.2-c-s1"
REF_DT = 0.1
DT_ORDER = [60, 20, 10, 6, 4, 2, 1, 0.6, 0.4, 0.3, 0.2, 0.1]
MODE_KO = {"bin": "빈평균", "pt": "점추출"}
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FILL = PatternFill("solid", fgColor="F2F2F2")
REF_FILL = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLUE = Font(color="1F60C4")     # 음수(비보수)
RED = Font(color="C00000")      # 양수(보수)


def load_a(path):
    """요약 CSV의 (a) 베어링 손상 섹션 → [dict]."""
    rows, insec, hdr = [], False, None
    for r in csv.reader(open(path, encoding="utf-8-sig")):
        if not r:
            continue
        c0 = r[0].lstrip("﻿")
        if c0.startswith("=== (a)"):
            insec = True; continue
        if c0.startswith("=== ("):     # 다른 섹션 시작
            insec = False; continue
        if not insec:
            continue
        if hdr is None and c0 == "bearing":
            hdr = r; continue
        if hdr and c0.startswith("Main"):
            rows.append(dict(zip(hdr, r)))
    return rows


def dt_of(path):
    tok = os.path.basename(path).split("_dt_")[1]
    return float(tok.split("_summary")[0].replace(".csv", ""))


def mode_of(path):
    return "bin" if "_bin_" in os.path.basename(path) else "pt"


def n_cases(dt):
    return 6001 if dt == REF_DT else int(round(600.0 / dt))


def collect():
    """{(mode,dt): {(bearing,std,basis): row}}"""
    data = {}
    for f in glob.glob(os.path.join("Results", f"T50_*_{DLC}_dt_*_summary.csv")):
        m, dt = mode_of(f), dt_of(f)
        for r in load_a(f):
            data.setdefault((m, dt), {})[(r["bearing"], r["standard"], r["life_basis"])] = r
    return data


def collect_aiso():
    """데이터 CSV에서 ISO16281 a_ISO = L10mr/L10r 점별 산출 → {(mode,dt,bearing): (평균, σ)}."""
    import statistics as st
    out = {}
    for f in glob.glob(os.path.join("Results", f"T50_*_{DLC}_dt_*.csv")):
        if f.endswith("_summary.csv"):
            continue
        m = mode_of(f)
        try:
            dt = dt_of(f)
        except (IndexError, ValueError):
            continue
        acc = {}
        for r in csv.DictReader(open(f, encoding="utf-8-sig")):
            try:
                l10r = float(r["L10r_rev"])
                if l10r > 0:
                    acc.setdefault(r["bearing"], []).append(float(r["L10mr_rev"]) / l10r)
            except (KeyError, ValueError):
                pass
        for b, vals in acc.items():
            if vals:
                out[(m, dt, b)] = (st.mean(vals), st.pstdev(vals) if len(vals) > 1 else 0.0)
    return out


def eps(cur, ref):
    try:
        c, rr = float(cur), float(ref)
        return (c - rr) / rr * 100 if rr else None
    except (TypeError, ValueError):
        return None


def style_eps(cell, v):
    if v is None:
        cell.value = "–"; return
    cell.value = v / 100.0
    cell.number_format = "+0.00%;-0.00%"
    cell.font = BLUE if v < 0 else RED


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_total(wb, data):
    ws = wb.create_sheet("총괄")
    cols = ["method", "dt_s", "n_cases", "bearing", "standard", "life_basis",
            "sample_damage", "scale_factor", "damage_30yr", "SF_fatigue",
            "equiv_life_years", "편향_ε_%(vs dt0.1)", "note"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c); cell.font = BOLD; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ref = data.get(("bin", REF_DT), {})       # dt=0.1 참값(두 방식 동일)
    keys = [("Main Bearing_UW", "ISO281", "basic"), ("Main Bearing_UW", "ISO281", "modified"),
            ("Main Bearing_UW", "ISO16281", "basic"), ("Main Bearing_UW", "ISO16281", "modified"),
            ("Main Bearing_DW", "ISO281", "basic"), ("Main Bearing_DW", "ISO281", "modified"),
            ("Main Bearing_DW", "ISO16281", "basic"), ("Main Bearing_DW", "ISO16281", "modified")]
    row = 2
    for mode in ("bin", "pt"):
        for dt in DT_ORDER:
            if mode == "pt" and dt == REF_DT:
                continue                       # 점추출 dt=0.1은 빈과 동일 → 생략
            block = data.get((mode, dt))
            if not block:
                continue
            for k in keys:
                r = block.get(k)
                if not r:
                    continue
                sd = r["sample_damage"]
                refrow = ref.get(k)
                ev = eps(sd, refrow["sample_damage"]) if refrow else None
                vals = [MODE_KO[mode], dt, n_cases(dt), k[0], k[1], k[2],
                        float(sd), int(float(r["scale_factor"])),
                        float(r["damage_30yr"]), float(r["SF_fatigue"]),
                        float(r["equiv_life_years"]), None, r.get("note", "")]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row, c)
                    if c == 12:
                        style_eps(cell, ev)
                    else:
                        cell.value = v
                    cell.border = BORDER
                ws.cell(row, 7).number_format = "0.000E+00"
                ws.cell(row, 9).number_format = "0.0000"
                ws.cell(row, 10).number_format = "0.000"
                ws.cell(row, 11).number_format = "#,##0.0"
                if dt == REF_DT:
                    for c in range(1, len(cols) + 1):
                        ws.cell(row, c).fill = REF_FILL
                row += 1
    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row-1}"
    autosize(ws, [9, 7, 8, 17, 10, 11, 13, 11, 12, 11, 15, 16, 16])
    return row - 2


def sheet_primary(wb, data):
    ws = wb.create_sheet("주지표요약")
    ws.append(["부록 8-5 · ISO16281 수정수명 (윤활유 50°C, DLC1.2-c-s1)"])
    ws.cell(1, 1).font = Font(bold=True, size=12)
    aiso = collect_aiso()
    hdr = ["dt (s)", "케이스수",
           "빈평균 UW수명(yr)", "점추출 UW수명(yr)", "빈평균 DW수명(yr)", "점추출 DW수명(yr)",
           "빈평균 UW 30년손상", "점추출 UW 30년손상", "빈평균 DW 30년손상", "점추출 DW 30년손상",
           "빈평균 UW a_ISO 평균", "빈평균 UW a_ISO σ", "점추출 UW a_ISO 평균", "점추출 UW a_ISO σ",
           "빈평균 DW a_ISO 평균", "빈평균 DW a_ISO σ", "점추출 DW a_ISO 평균", "점추출 DW a_ISO σ",
           "빈 편향 ε", "점 편향 ε"]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(2, c); cell.font = BOLD; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = BORDER

    ref = data.get(("bin", REF_DT), {})
    K = lambda b: (b, "ISO16281", "modified")
    refUW = ref.get(K("Main Bearing_UW"))
    r = 3
    for dt in DT_ORDER:
        def get(mode, b, fld):
            src = data.get(("bin", REF_DT) if dt == REF_DT else (mode, dt), {})
            rr = src.get(K(b))
            return float(rr[fld]) if rr else None

        ws.cell(r, 1, dt); ws.cell(r, 2, n_cases(dt))
        # 등가수명 (yr)
        for c, (mode, b) in zip((3, 4, 5, 6),
                                (("bin", "Main Bearing_UW"), ("pt", "Main Bearing_UW"),
                                 ("bin", "Main Bearing_DW"), ("pt", "Main Bearing_DW"))):
            ws.cell(r, c, get(mode, b, "equiv_life_years")); ws.cell(r, c).number_format = "#,##0.0"
        # 30년 손상
        for c, (mode, b) in zip((7, 8, 9, 10),
                                (("bin", "Main Bearing_UW"), ("pt", "Main Bearing_UW"),
                                 ("bin", "Main Bearing_DW"), ("pt", "Main Bearing_DW"))):
            ws.cell(r, c, get(mode, b, "damage_30yr")); ws.cell(r, c).number_format = "0.0000"
        # a_ISO 평균·σ (ISO16281, 열 11~18: 빈UW평균/σ, 점UW평균/σ, 빈DW평균/σ, 점DW평균/σ)
        for base, (mode, b) in zip((11, 13, 15, 17),
                                   (("bin", "Main Bearing_UW"), ("pt", "Main Bearing_UW"),
                                    ("bin", "Main Bearing_DW"), ("pt", "Main Bearing_DW"))):
            key = ("bin", REF_DT, b) if dt == REF_DT else (mode, dt, b)
            av = aiso.get(key)
            for j in range(2):
                cell = ws.cell(r, base + j)
                cell.value = av[j] if av else None
                cell.number_format = "0.0000"
        # 편향 (UW ISO16281 modified 표본손상 기준)
        for c, mode in ((19, "bin"), (20, "pt")):
            src = data.get(("bin", REF_DT) if dt == REF_DT else (mode, dt), {})
            rr = src.get(K("Main Bearing_UW"))
            ev = eps(rr["sample_damage"], refUW["sample_damage"]) if (rr and refUW) else None
            style_eps(ws.cell(r, c), ev)
        for c in range(1, len(hdr) + 1):
            ws.cell(r, c).border = BORDER
            if dt == REF_DT:
                ws.cell(r, c).fill = REF_FILL
        r += 1
    ws.freeze_panes = "C3"
    autosize(ws, [8, 10, 15, 15, 15, 15, 15, 15, 15, 15,
                  15, 13, 15, 13, 15, 13, 15, 13, 12, 12])

    # ── 하단: a_ISO 평균·σ 산출 절차 (개조식) ──
    mono = Font(name="Consolas", size=10)
    note_it = Font(italic=True, color="808080", size=9)
    r += 2
    ws.cell(r, 1, "■ a_ISO 평균·표준편차 산출 절차 (ISO 16281, 별도 해석 없이 저장 데이터로 산출)")
    ws.cell(r, 1).font = Font(bold=True, size=11)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="EAECEE")
    proc = [
        ("[1] 데이터원", None),
        ("    Results/T50_{bin|pt}_DLC1.2-c-s1_dt_<값>.csv  (점별/빈별 해석결과)", "m"),
        ("    사용 열: L10r_rev(기본 기준수명), L10mr_rev(수정 기준수명)", "m"),
        ("[2] 케이스별 a_ISO   (a1=1, 신뢰도 90%)", None),
        ("    a_ISO,i = L10mr_rev,i / L10r_rev,i     (L10r_rev,i > 0 인 케이스만; i=빈 또는 점)", "m"),
        ("[3] 조건별 집계   (조건 = 방식[빈/점] × dt × 베어링[UW/DW])", None),
        ("    n = 해당 조건의 케이스 수 (빈수 또는 점수)", "m"),
        ("[4] 평균", None),
        ("    a_bar = (1/n)·Σ a_ISO,i", "m"),
        ("[5] 표준편차 (모표준편차, ÷n)", None),
        ("    sigma = sqrt[ (1/n)·Σ (a_ISO,i − a_bar)^2 ]", "m"),
        ("※ dt=0.1은 빈=점 동일(빈당 1점) → 두 방식 값 동일", "i"),
        ("※ a_ISO = L10m/L10 이며 하중 P·점도비 κ의 함수. 50°C(κ=1.459)에서 a_ISO>1", "i"),
    ]
    for txt, sty in proc:
        r += 1
        c = ws.cell(r, 1, txt)
        if sty == "m":
            c.font = mono
        elif sty == "i":
            c.font = note_it


def main():
    data = collect()
    if not data:
        print("[오류] T50 요약 CSV를 찾지 못함")
        return
    wb = Workbook(); wb.remove(wb.active)
    n = sheet_total(wb, data)
    sheet_primary(wb, data)
    wb.save(OUT)
    print(f"[저장] {OUT}")
    print(f"  시트1 총괄: {n}행  ·  시트2 주지표요약: {len(DT_ORDER)}dt")
    print(f"  수집: {len(data)}개 (mode,dt) 조합")


if __name__ == "__main__":
    main()
