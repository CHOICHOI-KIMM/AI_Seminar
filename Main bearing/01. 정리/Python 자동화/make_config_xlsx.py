"""fatigue_config.xlsx (설정 템플릿) 생성/재생성. 실행: python make_config_xlsx.py"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "fatigue_config.xlsx")

ROWS = [
    ("MODEL_PATH",
     r"D:\AI\AI_Seminar\Main bearing\02. 자료\MASTA\26MW_메인베어링_기본설계_v1.3_샤프트 두께,형상 2안_피로하중 반영_260720.Masta",
     "대상 MASTA 모델(.Masta) 전체경로. v1.3형(Power Load 2개 + Load Case 1 필요)"),
    ("DLC_FILE",
     r"D:\AI\AI_Seminar\Main bearing\02. 자료\260714 유니슨 피로하중\DLC1.2-c-s1.$150",
     "시계열 하중 파일(.$150) 전체경로"),
    ("FATIGUE_HRS",
     r"D:\AI\AI_Seminar\Main bearing\02. 자료\260714 유니슨 피로하중\FatigueHours.txt",
     "발생빈도 파일(ScaleFactor·30년시간)"),
    ("DRIVER_LC", "Load Case 1",
     "드라이버 static 로드케이스명(Fatigue load case 내)"),
    ("DT_S", 60,
     "해석 dt[초]. 0.1=전량6001점 / 0.12=5001 / 0.15=4001 / 0.2=3001 / 0.4=1501 / 60=11. N=600/dt+1"),
    ("OUT_CSV", "",
     "출력 CSV '파일명'(비우면 자동: fatigue_<DLC>_dt<값>.csv). 전체경로를 넣으면 OUT_DIR보다 우선"),
    ("OUT_DIR", HERE,
     "결과 저장 폴더(CSV·요약·xlsx 저장 위치). 비우면 스크립트 폴더. 없으면 자동 생성"),
    ("MAKE_XLSX", "Y",
     "실행 후 엑셀(.xlsx, 데이터·요약 2시트) 자동생성 여부 Y/N"),
]

wb = Workbook()
ws = wb.active
ws.title = "설정"
ws.append(["파라미터", "값", "설명"])
for r in ROWS:
    ws.append(list(r))

hdr = PatternFill("solid", fgColor="D9E1F2")
for c in range(1, 4):
    cell = ws.cell(1, c)
    cell.font = Font(bold=True)
    cell.fill = hdr
val = PatternFill("solid", fgColor="FFF2CC")
for r in range(2, len(ROWS) + 2):
    ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2).fill = val          # 값 열 = 편집 대상(노란색)
    ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions[get_column_letter(1)].width = 16
ws.column_dimensions[get_column_letter(2)].width = 78
ws.column_dimensions[get_column_letter(3)].width = 60
ws.freeze_panes = "A2"

ws2 = wb.create_sheet("사용법")
for line in [
    "[사용 순서]",
    "1) '설정' 시트의 노란색 '값' 열만 수정 → 저장 → 엑셀 닫기",
    "2) run_fatigue.bat 더블클릭",
    "3) 완료 시 같은 폴더에 CSV 2개(데이터/요약) + .xlsx 생성",
    "",
    "[주의]",
    "- 실행 전 대상 .Masta 를 MASTA GUI에서 닫아주세요(lock 방지)",
    "- 출력 CSV/xlsx 를 엑셀에서 열어두면 쓰기 실패 → 닫아주세요",
    "- 중단돼도 다시 실행하면 남은 포인트부터 이어서 계산(체크포인트)",
    "",
    "[DT_S 참고]  N = 600/dt + 1",
    "  dt=60 → 11점(빠른시험) / dt=0.4 → 1501 / dt=0.2 → 3001 / dt=0.1 → 6001(전량, 약 30분)",
]:
    ws2.append([line])
ws2.column_dimensions["A"].width = 95

wb.save(OUT)
print("설정 엑셀 생성:", OUT)
