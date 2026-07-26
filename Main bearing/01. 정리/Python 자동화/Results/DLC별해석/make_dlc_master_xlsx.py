"""
DLC별해석_총괄.xlsx — 전 DLC 핵심 결과 + 최적 (dt, k)
데이터: dlc_master_summary.csv + dlc_meta.csv + DLC별 masta_best_summary.csv
- 손상·수명 절대값: MASTA 실측(참값 dt=0.1 + 최적조합 병기) 25개만 입력, 나머지 빈 칸
- 최적 (dt,k)·ε: 본해석 수행 DLC는 확정값·실측 ε(음영+비고), 그 외 스크리닝 예측
"""
import csv
import glob
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "DLC별해석_총괄.xlsx")
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL = PatternFill("solid", fgColor="FDEBD0")
MASTA_FILL = PatternFill("solid", fgColor="E2EFDA")     # MASTA 실측 행 음영
BOLD = Font(bold=True)
TH = Side(style="thin", color="BFBFBF")
BD = Border(left=TH, right=TH, top=TH, bottom=TH)
BLUE, RED = Font(color="1F60C4"), Font(color="C00000")
DESIGN_YEARS = 30.0
E_W = 9.0 / 8.0


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_best():
    """DLC별 masta_best_summary.csv → {DLC: dict} (참값+조합 산출값 포함)."""
    out = {}
    for p in glob.glob(os.path.join(HERE, "*", "masta_best_summary.csv")):
        r = list(csv.DictReader(open(p, encoding="utf-8-sig")))[0]
        if "D30_UW_ref" not in r:
            continue
        d = {k: float(v) for k, v in r.items()}
        # 조합값 유도: UW는 저장값, DW·Sys는 참값×(1+ε)
        d["D30_DW_cmb"] = d["D30_DW_ref"] * (1 + d["eps_DW"] / 100)
        d["life_UW_cmb"] = DESIGN_YEARS / d["D30_UW_cmb"]
        d["life_DW_cmb"] = DESIGN_YEARS / d["D30_DW_cmb"]
        d["life_Sys_cmb"] = d["life_Sys_ref"] / (1 + d["eps_Sys"] / 100)
        out[os.path.basename(os.path.dirname(p))] = d
    return out


def main():
    master = {r["DLC"]: r for r in csv.DictReader(
        open(os.path.join(HERE, "dlc_master_summary.csv"), encoding="utf-8-sig"))}
    meta = {r["DLC"]: r for r in csv.DictReader(
        open(os.path.join(HERE, "dlc_meta.csv"), encoding="utf-8-sig"))}
    best = load_best()

    wb = Workbook(); ws = wb.active; ws.title = "총괄"
    ws.append(["DLC별 해석 총괄 — 손상·수명 절대값은 MASTA 실측만 입력(음영 행, 참값 dt=0.1 + 최적조합 병기). "
               "스크리닝 절대값은 핀지지 절대편향 ≈3~5×로 미기재(부록 1). "
               "본해석 DLC의 최적 (dt,k)·ε = 확정값·실측, 그 외 = 스크리닝 예측"])
    ws.cell(1, 1).font = Font(bold=True, size=11, color="C00000")
    hdr = ["DLC", "ScaleFactor", "rpm 평균", "rpm CV%", "T_3P [s]", "dt 상한(5/rpm_max)",
           "M_X CV%", "저속점(rpm<1)", "κ클립 점수",
           "UW 30년손상(참값)", "DW 30년손상(참값)",
           "UW 수명(참값,yr)", "DW 수명(참값,yr)", "Sys 수명(참값,yr)",
           "UW 30년손상(조합)", "DW 30년손상(조합)",
           "UW 수명(조합,yr)", "DW 수명(조합,yr)", "Sys 수명(조합,yr)",
           "최적 dt [s]", "최적 k", "ε_UW@best", "ε_DW@best", "ε_Sys@best", "비고"]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(2, c); cell.font = BOLD; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = BD

    C_ABS = list(range(10, 20))          # 절대값 10열
    C_EPS = (22, 23, 24)
    r = 3
    for name in sorted(master):
        m, mt = master[name], meta.get(name, {})
        b = best.get(name)
        valid = m.get("valid") == "1"
        flags = []
        if int(float(m.get("n_rpm_lt1", 0) or 0)) > 0:
            flags.append("저속")
        if int(float(m.get("n_kappa_clip", 0) or 0)) > 0:
            flags.append("κ클립")
        if not valid:
            flags.append("참값손상≈0(스킵)")
        if valid and m.get("best_dt_s") in ("", None, "None"):
            flags.append("합격조합없음")
        if b:
            tag = "MASTA 실측"
            if b["ncorr"] > 0:
                tag += f"(보정{int(b['ncorr'])}회)"
            flags.insert(0, tag)
        # 절대값 (MASTA 실측만) + dt·k·ε (실측 우선)
        if b:
            absv = [b["D30_UW_ref"], b["D30_DW_ref"],
                    b["life_UW_ref"], b["life_DW_ref"], b["life_Sys_ref"],
                    b["D30_UW_cmb"], b["D30_DW_cmb"],
                    b["life_UW_cmb"], b["life_DW_cmb"], b["life_Sys_cmb"]]
            dtk = [b["dt"], b["k"], b["eps_UW"], b["eps_DW"], b["eps_Sys"]]
        else:
            absv = [None] * 10
            dtk = [fnum(m.get("best_dt_s")), fnum(m.get("best_k")),
                   fnum(m.get("eps_UW_at_best")), fnum(m.get("eps_DW_at_best")),
                   fnum(m.get("eps_Sys_at_best"))]
        vals = [name, fnum(m["ScaleFactor"]), fnum(m["rpm_mean"]), fnum(m["rpm_CV_pct"]),
                fnum(m["T3P_s"]), fnum(m["dt_rule_max_s"]),
                fnum(mt.get("Mz_CV_pct")), int(float(m.get("n_rpm_lt1", 0) or 0)),
                int(float(m.get("n_kappa_clip", 0) or 0)),
                *absv, *dtk, ", ".join(flags)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v if v is not None else "–")
            cell.border = BD
            if c in C_EPS and isinstance(v, float):
                cell.value = v / 100.0
                cell.number_format = "+0.00%;-0.00%"
                cell.font = BLUE if v < 0 else RED
            if b and c in C_ABS + [20, 21, 22, 23, 24]:
                cell.fill = MASTA_FILL
        for c, fmt in ((2, "#,##0"), (3, "0.000"), (4, "0.0"), (5, "0.00"), (6, "0.00"),
                       (7, "0.0"), (10, "0.0000"), (11, "0.0000"),
                       (12, "#,##0.0"), (13, "#,##0.0"), (14, "#,##0.0"),
                       (15, "0.0000"), (16, "0.0000"),
                       (17, "#,##0.0"), (18, "#,##0.0"), (19, "#,##0.0"),
                       (20, "0.0"), (21, "0.00")):
            ws.cell(r, c).number_format = fmt
        if flags and not b:
            ws.cell(r, 25).fill = WARN_FILL
        r += 1
    # ── 전 DLC 합산 (MASTA 기반 — 부분합/전량) ──
    r += 1
    ws.cell(r, 1, "■ 전 DLC 합산 (MASTA 기반)")
    ws.cell(r, 1).font = Font(bold=True, size=11)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="EAECEE")
    part = None
    pp = os.path.join(HERE, "fleet_masta_partial.csv")
    if os.path.exists(pp):
        part = list(csv.DictReader(open(pp, encoding="utf-8-sig")))[0]
    lbl = (f"Σ D30 (상위 {part['n_dlc']} DLC 부분합, MASTA 참값)" if part
           else "Σ D30 (전 111 DLC, MASTA)")
    r += 1
    ws.cell(r, 1, lbl).font = BOLD
    for c, key in ((10, "sumD30_UW"), (11, "sumD30_DW")):
        cell = ws.cell(r, c, float(part[key]) if part else "–")
        cell.border = BD
        if part:
            cell.number_format = "0.0000"
    ws.cell(r, 1).border = BD
    r += 1
    ws.cell(r, 1, "수명 [yr] = 30/ΣD30 · Sys=와이블(e=9/8)"
            + (" — 부분합 기준" if part else "")).font = BOLD
    for c, key in ((12, "life_UW"), (13, "life_DW"), (14, "life_Sys")):
        cell = ws.cell(r, c, float(part[key]) if part else "–")
        cell.border = BD
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        if part:
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=True, color="C00000")
    ws.cell(r, 1).border = BD

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr))}{r-1}"
    for i, w in enumerate([15, 11, 9, 8, 9, 10, 8, 9, 9,
                           12, 12, 12, 12, 12, 12, 12, 12, 12, 12,
                           9, 8, 10, 10, 10, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(OUT)
    print(f"[저장] {OUT}  (MASTA 실측 {sum(1 for _ in best)}개 DLC 입력)")


if __name__ == "__main__":
    main()
