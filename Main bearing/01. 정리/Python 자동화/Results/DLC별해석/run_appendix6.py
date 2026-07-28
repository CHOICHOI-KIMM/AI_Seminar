"""
부록 6 — LDD(하중지속분포) 기반 등가하중 단일 로드케이스 피로수명 (v1.3 50°C)
============================================================================
DesignLoads-B 'FatigueLoads-LDD'의 6 DOF 중 토크(Mx) 제외 5 DOF(My·Mz·Fx·Fy·Fz)를
각각 time-ratio 가중 λ=10/3 등가하중으로 축약(부호=시간가중 평균 부호) → 좌표변환 →
1개 MASTA 로드케이스 입력 → 베어링 L10mr → 30년 손상(D30=총회전수/L10mr) → 수명.
부록 1~4 참값 함대수명(UW 4.44 / Sys 3.93 yr)과 대비.
"""
import csv
import math
import os
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, ROOT)

MODEL = (r"D:\AI\AI_Seminar\Main bearing\02. 자료\MASTA"
         r"\26MW_메인베어링_기본설계_v1.3_샤프트 두께,형상 2안"
         r"_피로하중 반영_유연체_FE_온도_50도_260721.Masta")
LDDXLSX = (r"D:\AI\AI_Seminar\Main bearing\02. 자료\260714 유니슨 피로하중"
           r"\EX.20260713.K26.DE.MainBearing_DesignLoads-B.xlsx")
OUTDIR = os.path.join(HERE, "부록6_LDD")
LAM = 10.0 / 3.0
DESIGN_YEARS = 30.0
DOFS = ["Mx", "My", "Mz", "Fx", "Fy", "Fz"]      # 파일 좌표


def equiv_loads():
    """LDD → DOF별 (등가크기, 시간가중평균부호적용값, 총시간)."""
    ws = load_workbook(LDDXLSX, read_only=True, data_only=True)["FatigueLoads-LDD"]
    rows = list(ws.iter_rows(values_only=True))[2:]
    out = {}
    for j, d in enumerate(DOFS):
        Lc, Tc = 2 * j, 2 * j + 1
        data = [(float(r[Lc]), float(r[Tc])) for r in rows
                if r[Lc] is not None and r[Tc] is not None]
        T = sum(t for _, t in data)
        if T <= 0:
            out[d] = (0.0, 0.0, 0.0)
            continue
        Leq = (sum(t * abs(L) ** LAM for L, t in data) / T) ** (1.0 / LAM)
        wmean = sum(t * L for L, t in data) / T
        signed = math.copysign(Leq, wmean) if wmean != 0 else Leq
        out[d] = (Leq, signed, T)
    return out


def main():
    import masta_fatigue as mf
    from mastapy.system_model import Design
    from mastapy.system_model.analyses_and_results.static_loads import AnalysisType

    os.makedirs(OUTDIR, exist_ok=True)
    tot_rev, mean_rpm, tot_h = [float(x) for x in
                               open(os.path.join(HERE, "appendix6_speed.txt")).read().split()]
    eq = equiv_loads()
    print("[등가하중] DOF별 (크기 / 부호적용):")
    for d in DOFS:
        L, s, T = eq[d]
        tag = " ← 토크·제외" if d == "Mx" else ""
        print(f"  {d}: |{L:,.1f}|  → {s:+,.1f}{tag}")

    # 파일좌표 rec (토크 Mx=0으로 제외), 부호적용 등가값
    rec = {"rpm": mean_rpm, "Mx": 0.0,
           "My": eq["My"][1], "Mz": eq["Mz"][1],
           "Fx": eq["Fx"][1], "Fy": eq["Fy"][1], "Fz": eq["Fz"][1]}
    print(f"[속도] 피로 평균 rpm = {mean_rpm:.4f} · 30년 총회전 {tot_rev:.4e} Rev")

    design = Design.load(MODEL)
    asm = design.all_parts_of_type_root_assembly()[0]
    pl = list(asm.all_parts_of_type_point_load())[0]
    ipl = next(p for p in asm.all_parts_of_type_power_load() if "input" in str(p).lower())
    bearings = {("UW" if "UW" in str(b) else "DW"): b
                for b in asm.all_parts_of_type_bearing()}
    lc = next(c for c in asm.design_properties.static_loads if c.name == "Load Case 1")
    print("[모델] v1.3 50°C 로드 완료")

    loads = mf.set_loads(lc, pl, ipl, rec)
    print("[MASTA 입력]", {k: round(v, 1) for k, v in loads.items()})
    sd = lc.analysis_of(AnalysisType.SYSTEM_DEFLECTION)
    sd.perform_analysis()

    res = {}
    for key, b in bearings.items():
        d = sd.results_for(b).component_detailed_analysis
        l10mr = float(d.isots162812008.modified_reference_rating_life_cycles)
        l10 = float(d.iso2812007.basic_rating_life_cycles)
        P = float(mf.g(d, "iso2812007.dynamic_equivalent_load") or 0)
        d30 = tot_rev / l10mr if l10mr > 0 else float("inf")
        life = DESIGN_YEARS / d30 if d30 > 0 else float("inf")
        res[key] = dict(P=P, l10=l10, l10mr=l10mr, d30=d30, life=life)
        print(f"  [{key}] P={P/1e3:,.0f}kN L10={l10:.3e} L10mr={l10mr:.3e} "
              f"D30={d30:.4f} 수명={life:,.1f}yr")

    lU, lD = res["UW"]["life"], res["DW"]["life"]
    lS = (lU ** (-9 / 8) + lD ** (-9 / 8)) ** (-8 / 9)
    print(f"  [Sys] {lS:.2f} yr (와이블 e=9/8)")

    with open(os.path.join(OUTDIR, "appendix6_result.csv"), "w", newline="",
              encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["DOF", "Leq_abs", "Leq_signed", "unit"])
        for d in DOFS:
            u = "kNm" if d.startswith("M") else "kN"
            w.writerow([d, eq[d][0], eq[d][1], u + (" (제외)" if d == "Mx" else "")])
        w.writerow([])
        w.writerow(["mean_rpm", mean_rpm, "tot_rev_30yr", tot_rev, "tot_h", tot_h])
        w.writerow([])
        w.writerow(["bearing", "P_N", "L10_rev", "L10mr_rev", "D30", "life_yr"])
        for key in ("UW", "DW"):
            r = res[key]
            w.writerow([key, r["P"], r["l10"], r["l10mr"], r["d30"], r["life"]])
        w.writerow(["Sys", "", "", "", "", lS])
    print(f"\n[완료] 부록6_LDD/appendix6_result.csv")
    print(f"  LDD 등가법: UW {lU:.2f} / DW {lD:.2f} / Sys {lS:.2f} yr")
    print(f"  부록1~4 참값 함대: UW 4.44 / DW 24.10 / Sys 3.93 yr")


if __name__ == "__main__":
    main()
