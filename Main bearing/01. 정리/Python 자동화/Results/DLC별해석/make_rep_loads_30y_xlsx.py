"""
30년기준 대표하중 111 DLC 통합 xlsx
====================================
대표하중_30년기준/ 의 {DLC}_30y.csv 111개를 단일 시트로 통합한다.

  · A열 DLC · B열 index(전체 통산 1~2646) · 이후 rpm, rev_30y, 하중 6열
  · 개별 CSV 는 수정하지 않음 (index 0-based 유지)
"""
import csv
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "대표하중_원본좌표", "대표하중_30년기준")
OUT = os.path.join(SRC, "대표하중_30년기준_통합.xlsx")

HEAD = ["DLC", "index", "rpm", "rev_30y",
        "Fx [kN]", "Fy [kN]", "Fz [kN]", "Mx [kNm]", "My [kNm]", "Mz [kNm]"]
NUMFMT = {"rpm": "0.00000", "rev_30y": "#,##0.000"}
WIDTH = {"DLC": 20, "index": 8}


def main():
    names = sorted(f[:-8] for f in os.listdir(SRC) if f.endswith("_30y.csv"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "대표하중_30년기준"
    ws.append(HEAD)

    n = 0
    for nm in names:
        with open(os.path.join(SRC, f"{nm}_30y.csv"), encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                n += 1
                ws.append([nm, n] + [float(r[c]) for c in HEAD[2:]])
    last = ws.max_row

    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="44546A")
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill = hf, fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        L = get_column_letter(c)
        ws.column_dimensions[L].width = WIDTH.get(HEAD[c - 1], 13)
        fmt = NUMFMT.get(HEAD[c - 1], "#,##0.000" if c > 4 else None)
        if fmt:
            for row in range(2, last + 1):
                ws.cell(row, c).number_format = fmt
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{last}"

    wb.save(OUT)
    print(f"[완료] {len(names)}개 DLC · {n:,}행 → {OUT}")
    print(f"       크기 {os.path.getsize(OUT)/1024:.0f} KB")


if __name__ == "__main__":
    main()
