"""
부록 5 후처리 — DLC별 집계 + 엑셀 생성
========================================
lub_per_bin.csv (15,876행) →

  lub_per_dlc.csv        모델 × DLC × 베어링 별 회전수 가중평균 + 단순평균
  윤활조건_검토.xlsx     요약 1시트 + 모델별 상세 3시트

가중평균은 **30년 회전수**(rev × ScaleFactor) 가중이다 — 손상가중과 기준이 같아야
비교가 성립한다. 단순평균은 **빈 단위** 동등가중이다. (260731 정정: 이전에는
1회 시뮬레이션 rev 로 가중해 ScaleFactor 3자릿수 차이를 반영하지 못했다.)
윤활 체제는 λ 로 구분한다 — 경계(<1) · 혼합(1~3) · 유체(>3).

rpm 만은 **30년 운전시간**(Δt × ScaleFactor) 가중이다. rev 가 rpm 에 정비례해
(`rev = |n|/60 · Δt`) 회전수로 가중하면 Σn²/Σn 즉 대조화평균이 되어 최대 +56%
부풀기 때문이다(260731 정정). 빈 Δt 는 rev·60/|rpm| 로 복원한다.
"""
import collections
import csv
import os
import statistics as st

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DIR = os.path.join(HERE, "부록5_윤활조건")
SRC = os.path.join(DIR, "lub_per_bin.csv")
MODELS = ["기준선", "A1", "B1"]
# DLC별 손상 — 기준선은 P2 Phase 1, A1·B1 은 Phase 2
DMG_SRC = [(os.path.join(HERE, "P2_피로수명_Phase1", "fatigue_per_dlc.csv"),
            {"base": "기준선"}),
           (os.path.join(HERE, "P2_피로수명_Phase2", "fatigue_per_dlc.csv"),
            {"A1": "A1", "B1": "B1"})]
VARS = ["kappa", "lambda_in", "lambda_out", "hmin_in_um", "hmin_out_um",
        "nu_mm2s", "nu1_mm2s", "a_iso"]


def f(r, k):
    try:
        return float(r[k])
    except (TypeError, ValueError):
        return None


def regime(lam):
    return "경계 (λ<1)" if lam < 1 else ("혼합 (1≤λ<3)" if lam < 3 else "유체 (λ≥3)")


def load_damage():
    d = {}
    for path, mp in DMG_SRC:
        if not os.path.isfile(path):
            continue
        with open(path, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                if r["design"] in mp:
                    d[(mp[r["design"]], r["DLC"])] = r
    return d


def main():
    with open(SRC, encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    dmg = load_damage()
    tot = collections.defaultdict(float)
    for (m, _), r in dmg.items():
        tot[m] += float(r["D30_UW"])
    for r in rows:
        sf = f(r, "ScaleFactor") or 1.0
        rev, rpm = f(r, "rev") or 0.0, f(r, "rpm") or 0.0
        r["_rev"] = rev * sf
        r["_dur"] = (rev * 60.0 / abs(rpm) if rpm else 20.0) * sf
        for v in VARS:
            r[v] = f(r, v)
        r["_cap"] = int(r["a_iso_capped"] or 0)

    # ── DLC별 집계 ──
    grp = collections.defaultdict(list)
    for r in rows:
        grp[(r["model"], r["DLC"], r["brg"])].append(r)
    out = []
    for (m, d, b), g in grp.items():
        W = sum(x["_rev"] for x in g)
        row = dict(model=m, DLC=d, brg=b, nbin=len(g),
                   k=g[0]["k"], ScaleFactor=g[0]["ScaleFactor"],
                   rev_total=round(W, 4),
                   rpm_tw=round(sum(f(x, "rpm") * x["_dur"] for x in g)
                                / sum(x["_dur"] for x in g), 4),
                   dur_total=round(sum(x["_dur"] for x in g), 1))
        row["nbin_all"] = len(g)
        for v in VARS:
            vals = [(x[v], x["_rev"]) for x in g if x[v] is not None]
            if not vals:
                row[f"{v}_w"] = row[f"{v}_m"] = ""
                continue
            ww = sum(w for _, w in vals)
            row[f"{v}_w"] = round(sum(a * w for a, w in vals) / ww, 5) if ww else ""
            row[f"{v}_m"] = round(st.fmean([a for a, _ in vals]), 5)
        row["cap_pct"] = round(100.0 * sum(x["_cap"] for x in g) / len(g), 1)
        lam = [x["lambda_in"] for x in g if x["lambda_in"] is not None]
        row["lam_min"] = round(min(lam), 4) if lam else ""
        row["lam_max"] = round(max(lam), 4) if lam else ""
        row["regime_w"] = regime(row["lambda_in_w"]) if row["lambda_in_w"] != "" else ""
        dd = dmg.get((m, d))
        row["D30_UW"] = round(float(dd["D30_UW"]), 6) if dd else ""
        row["D30_DW"] = round(float(dd["D30_DW"]), 6) if dd else ""
        row["dmg_pct"] = (round(100.0 * float(dd["D30_UW"]) / tot[m], 3)
                          if dd and tot[m] > 0 else "")
        out.append(row)
    out.sort(key=lambda r: (MODELS.index(r["model"]), r["DLC"], r["brg"]))
    p = os.path.join(DIR, "lub_per_dlc.csv")
    with open(p, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=list(out[0]))
        w.writeheader(); w.writerows(out)
    print(f"[저장] lub_per_dlc.csv — {len(out)}행")

    # ── 엑셀 ──
    wb = openpyxl.Workbook()
    hf, fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="44546A")

    def style(ws, nhdr=1, widths=None):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(nhdr, c)
            cell.font, cell.fill = hf, fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = (
                (widths or {}).get(ws.cell(nhdr, c).value, 13))
        ws.freeze_panes = ws.cell(nhdr + 1, 4)
        ws.auto_filter.ref = (f"A{nhdr}:"
                              f"{get_column_letter(ws.max_column)}{ws.max_row}")

    # 요약
    ws = wb.active
    ws.title = "요약"
    ws.append(["모델", "베어링", "DLC수", "30년 회전수",
               "κ 30년가중", "κ 빈단순", "λ_in 30년가중", "λ_in 빈단순", "λ_out 30년가중",
               "h_min,in 30년가중 [µm]", "a_ISO 30년가중", "a_ISO 빈단순", "a_ISO 손상가중",
               "a_ISO<1 DLC", "포화 빈 [%]", "경계 12건 손상 [%]",
               "경계 λ<1", "혼합 1≤λ<3", "유체 λ≥3"])
    for m in MODELS:
        for b in ("UW", "DW"):
            g = [r for r in out if r["model"] == m and r["brg"] == b]
            W = sum(r["rev_total"] for r in g)
            def wm(key):
                return sum(r[key] * r["rev_total"] for r in g if r[key] != "") / W
            reg = collections.Counter(r["regime_w"] for r in g)
            ws.append([m, b, len(g), round(W, 1),
                       round(wm("kappa_w"), 4), round(sum(r["kappa_m"]*r["nbin_all"] for r in g)/sum(r["nbin_all"] for r in g), 4),
                       round(wm("lambda_in_w"), 4),
                       round(sum(r["lambda_in_m"]*r["nbin_all"] for r in g)/sum(r["nbin_all"] for r in g), 4),
                       round(wm("lambda_out_w"), 4),
                       round(wm("hmin_in_um_w"), 4),
                       round(wm("a_iso_w"), 4),
                       round(sum(r["a_iso_m"]*r["nbin_all"] for r in g)/sum(r["nbin_all"] for r in g), 4),
                       (round(sum(r["a_iso_w"] * r["D30_UW"] for r in g
                                  if r["D30_UW"] != "") /
                              sum(r["D30_UW"] for r in g if r["D30_UW"] != ""), 4)
                        if any(r["D30_UW"] != "" for r in g) else ""),
                       sum(1 for r in g if r["a_iso_w"] != "" and r["a_iso_w"] < 1),
                       round(st.fmean([r["cap_pct"] for r in g]), 2),
                       (round(100.0 * sum(r["D30_UW"] for r in g
                                          if r["D30_UW"] != "" and r["lambda_in_w"] < 1) /
                              sum(r["D30_UW"] for r in g if r["D30_UW"] != ""), 3)
                        if any(r["D30_UW"] != "" for r in g) else ""),
                       reg.get("경계 (λ<1)", 0), reg.get("혼합 (1≤λ<3)", 0),
                       reg.get("유체 (λ≥3)", 0)])
    style(ws, widths={"모델": 10, "DLC수": 8, "베어링": 9})

    # 모델별 상세
    HD = ["DLC", "베어링", "빈수", "k", "rpm 30년 시간가중", "30년 운전시간 [s]",
          "30년 회전수",
          "κ 30년가중", "κ 빈단순", "λ_in 30년가중", "λ_in 빈단순", "λ_in 최소", "λ_in 최대",
          "λ_out 30년가중", "h_min,in [µm]", "h_min,out [µm]",
          "ν [mm²/s]", "ν₁ 30년가중 [mm²/s]", "a_ISO 30년가중", "a_ISO 빈단순",
          "포화 [%]", "윤활 체제", "D30_UW", "D30_DW", "손상 기여 [%]"]
    KEYS = ["DLC", "brg", "nbin", "k", "rpm_tw", "dur_total", "rev_total",
            "kappa_w", "kappa_m", "lambda_in_w", "lambda_in_m", "lam_min", "lam_max",
            "lambda_out_w", "hmin_in_um_w", "hmin_out_um_w",
            "nu_mm2s_w", "nu1_mm2s_w", "a_iso_w", "a_iso_m", "cap_pct", "regime_w",
            "D30_UW", "D30_DW", "dmg_pct"]
    for m in MODELS:
        w2 = wb.create_sheet(m)
        w2.append(HD)
        for r in [x for x in out if x["model"] == m]:
            w2.append([r[k] for k in KEYS])
        style(w2, widths={"DLC": 20, "윤활 체제": 15, "베어링": 9, "빈수": 7})
    xp = os.path.join(DIR, "윤활조건_검토.xlsx")
    wb.save(xp)
    print(f"[저장] {os.path.basename(xp)} — 요약 + {len(MODELS)}시트 "
          f"({os.path.getsize(xp)/1024:.0f} KB)")


if __name__ == "__main__":
    main()
