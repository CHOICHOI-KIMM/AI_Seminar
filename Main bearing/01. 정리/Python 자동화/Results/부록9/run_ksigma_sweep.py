"""
부록 9 · 2단계 MASTA 검증 스윕 (무인, 배치 모드 N=20)
=====================================================
k ∈ {0.15, 0.20, 0.25} × dt ∈ {10, 4, 2, 1, 0.6}  (2,110빈/k, 총 6,330해석)
- 대표하중: 5분력 μ+sign(μ)·k·σ (파일좌표, 변환은 ±1 배율이라 MASTA좌표와 동치) · M_Z·rpm 산술평균
- 배치: duplicate(ds, name) × N=20 → 전용 duty cycle → component_analysis_cases 추출 (§10-6.1)
- 참값: 부록 8 T50_bin dt=0.1 요약 재사용
- (k,dt) 완료마다: 데이터/요약 CSV + xlsx(수식 오인 수정) + §9-7 표 + timing CSV 갱신
- 자체검증: 최초 1빈을 단일 경로로도 해석해 배치와 대조
"""
import csv
import math
import os
import sys
import time

import psutil

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, ROOT)


MODEL = (r"D:\AI\AI_Seminar\Main bearing\02. 자료\MASTA"
         r"\26MW_메인베어링_기본설계_v1.3_샤프트 두께,형상 2안"
         r"_피로하중 반영_유연체_FE_온도_50도_260721.Masta")
KS = [0.15, 0.20, 0.25]
DTS = [20, 10, 4, 2, 1, 0.6]
NBATCH = 20
DT0 = 0.1
MEM_LIMIT = 90.0
DLC_NAME = "DLC1.2-c-s1"
DOC = os.path.join(ROOT, "DLC기반_피로해석_프로세스_v1.md")
MS, ME = "<!-- APPENDIX_9_S2_START -->", "<!-- APPENDIX_9_S2_END -->"
TIMING_CSV = os.path.join(HERE, "ksigma_timing.csv")
REF_SUMMARY = os.path.join(ROOT, "Results", "부록8", "T50_bin_DLC1.2-c-s1_dt_0.1_summary.csv")
EPS_MAP = os.path.join(HERE, "ksigma_screen_eps_map.csv")
PRIMARY = ("Main Bearing_UW", "ISO16281", "modified")

# 5분력(+kσ 대상, 파일좌표) / 산술평균 유지 성분
KSIG_KEYS = ("Fz", "Fy", "Fx", "Mz", "My")
MEAN_KEYS = ("rpm", "Mx")


def ktag(k):
    return ("%g" % k)


def paths(k, dt):
    b = os.path.join(HERE, f"T50_ks{ktag(k)}_{DLC_NAME}_dt_{'%g' % dt}")
    return b + ".csv", b + "_summary.csv"


# ── 참값·예측 로드 ──
def load_ref():
    for r in csv.reader(open(REF_SUMMARY, encoding="utf-8-sig")):
        if len(r) > 8 and (r[0], r[1], r[2]) == PRIMARY:
            return float(r[3])
    raise RuntimeError("참값 요약에서 주지표를 찾지 못함")


def load_pred():
    out = {}
    for r in csv.DictReader(open(EPS_MAP, encoding="utf-8-sig")):
        out[(float(r["k"]), float(r["dt_s"]))] = float(r["eps_UW_pct"])
    return out


# ── 빈 대표하중 ──
def bin_reps(data, dt, k):
    """→ [(bin_index, t0, rpm_avg, rev, rec)]"""
    kp = int(round(dt / DT0))
    n = len(data)
    nb = n // kp
    edges = [(b * kp, (b + 1) * kp) for b in range(nb)]
    if edges[-1][1] < n:
        edges[-1] = (edges[-1][0], n)
    out = []
    for bi, (i0, i1) in enumerate(edges):
        m = i1 - i0
        span = m * DT0
        rec = {}
        for key in MEAN_KEYS:
            rec[key] = sum(data[i][key] for i in range(i0, i1)) / m
        for key in KSIG_KEYS:
            mu = sum(data[i][key] for i in range(i0, i1)) / m
            var = sum((data[i][key] - mu) ** 2 for i in range(i0, i1)) / m
            rec[key] = mu + math.copysign(1.0, mu) * k * math.sqrt(var)
        rec["t"] = round(i0 * DT0, 4)
        out.append((bi, rec["t"], rec["rpm"], rec["rpm"] / 60.0 * span, rec))
    return out


# ── 결과 수집 → ε ──
def eps_of(summary_csv, ref):
    for r in csv.reader(open(summary_csv, encoding="utf-8-sig")):
        if len(r) > 8 and (r[0], r[1], r[2]) == PRIMARY:
            return (float(r[3]) / ref - 1) * 100
    return None


DONE = {}      # {(k,dt): (eps, ms_per_case)}  (콘솔 로그용)
N_BINS = {20: 30, 10: 60, 4: 150, 2: 300, 1: 600, 0.6: 1000}


def _sd(path, bearing):
    """→ (표본손상, 등가수명 yr) — ISO16281 modified."""
    for r in csv.reader(open(path, encoding="utf-8-sig")):
        if len(r) > 8 and r[0] == bearing and r[1] == "ISO16281" and r[2] == "modified":
            return float(r[3]), float(r[7])
    return None


def update_doc(note=""):
    """무상태 갱신: 요약 CSV(UW·DW) + 스크리닝 맵 + timing CSV 에서 §9-7a/b/c 재구성."""
    # 참값 (UW·DW)
    ref = {b: _sd(REF_SUMMARY, b) for b in ("Main Bearing_UW", "Main Bearing_DW")}
    ref_life = {b: (ref[b][1] if ref[b] else None) for b in ref}
    # 스크리닝 예측 (UW·DW)
    pred = {}
    if os.path.exists(EPS_MAP):
        for r in csv.DictReader(open(EPS_MAP, encoding="utf-8-sig")):
            pred[(float(r["k"]), float(r["dt_s"]))] = (
                float(r["eps_UW_pct"]), float(r["eps_DW_pct"]))
    # 타이밍 (ms/케이스)
    tim = {}
    if os.path.exists(TIMING_CSV):
        for r in csv.DictReader(open(TIMING_CSV, encoding="utf-8-sig")):
            try:
                tim[(float(r["k"]), float(r["dt_s"]))] = float(r["ms_per_case"])
            except ValueError:
                pass

    def eps_cell(k, dt, bidx, bearing):
        pr = pred.get((k, float(dt)))
        prs = f"{pr[bidx]:+.2f}" if pr else "?"
        _, sc = paths(k, dt)
        if os.path.exists(sc) and ref[bearing]:
            v = _sd(sc, bearing)
            if v is None:
                return f"–({prs})"
            e = (v[0] / ref[bearing][0] - 1) * 100
            mark = "✅" if 0.0 <= e <= 3.0 else "❌"
            return f"**{e:+.2f}%**({prs}) {mark} · {v[1]:,.1f} yr"
        return f"–({prs})"

    L = [MS, "",
         f"> 배치 N={NBATCH} (§10-6.3) · 참값 = 부록 8 T50_bin dt=0.1 (베어링별) · "
         f"합격 0 ≤ ε ≤ +3% (베어링별 판정) · 괄호 안 = 1단계 스크리닝 예측. "
         f"UW = 파손 지배 주지표, DW 병기(260723 추가) · 셀 = ε 실측(예측) 판정 · **등가수명**. "
         f"참값 등가수명: UW {ref_life['Main Bearing_UW']:,.1f} yr / "
         f"DW {ref_life['Main Bearing_DW']:,.1f} yr. {note}", ""]
    for title, bidx, bearing in (("#### 9-7a. UW (주지표)", 0, "Main Bearing_UW"),
                                 ("#### 9-7b. DW", 1, "Main Bearing_DW")):
        L += [title, "",
              "| dt (s) | 빈수 | k=0.15 ε 실측(예측) | k=0.20 ε 실측(예측) | k=0.25 ε 실측(예측) |",
              "|-------:|-----:|-------------------:|-------------------:|-------------------:|"]
        for dt in DTS:
            cells = [eps_cell(k, dt, bidx, bearing) for k in KS]
            L.append(f"| {dt:g} | {N_BINS[dt]:,} | {cells[0]} | {cells[1]} | {cells[2]} |")
        L.append("")
    # 9-7c. 시스템 수명 (직렬 와이블 조합, e=9/8)
    E_W = 9.0 / 8.0

    def lsys_of(lu, ld):
        return (lu ** -E_W + ld ** -E_W) ** (-1.0 / E_W)

    ref_ls = (lsys_of(ref["Main Bearing_UW"][1], ref["Main Bearing_DW"][1])
              if all(ref.values()) else None)
    L += ["#### 9-7c. 시스템 수명 L10,sys (직렬 와이블, e=9/8)", "",
          f"> `L10,sys = [L_UW^(-9/8) + L_DW^(-9/8)]^(-8/9)` (Lundberg-Palmgren 직렬 조합, "
          f"독립·동일기울기 가정) · ε = 참값L10,sys/L10,sys − 1 (손상 관점, +=보수측, 합격 0~+3%) · "
          f"**참값 L10,sys = {ref_ls:,.1f} yr** "
          f"(UW {ref['Main Bearing_UW'][1]:,.1f} / DW {ref['Main Bearing_DW'][1]:,.1f})", "",
          "| dt (s) | 빈수 | k=0.15 [yr] | k=0.20 [yr] | k=0.25 [yr] |",
          "|-------:|-----:|-----------:|-----------:|-----------:|"]
    for dt in DTS:
        cells = []
        for k in KS:
            _, sc = paths(k, dt)
            if os.path.exists(sc):
                vu, vd = _sd(sc, "Main Bearing_UW"), _sd(sc, "Main Bearing_DW")
                if vu and vd:
                    ls = lsys_of(vu[1], vd[1])
                    es = (ref_ls / ls - 1) * 100      # 손상 관점 ε (+=보수측)
                    mark = "✅" if 0.0 <= es <= 3.0 else "❌"
                    cells.append(f"{ls:,.1f} yr · **{es:+.2f}%** {mark}")
                else:
                    cells.append("–")
            else:
                cells.append("–")
        L.append(f"| {dt:g} | {N_BINS[dt]:,} | {cells[0]} | {cells[1]} | {cells[2]} |")
    L.append("")

    # 9-7d. 해석시간 (dt/k별)
    L += ["#### 9-7d. 해석시간 (dt/k별)", "",
          "| dt (s) | 빈수 | k=0.15 [s] (ms/케이스) | k=0.20 [s] (ms/케이스) | k=0.25 [s] (ms/케이스) |",
          "|-------:|-----:|----------------------:|----------------------:|----------------------:|"]
    for dt in DTS:
        nb = N_BINS[dt]
        cells = []
        for k in KS:
            ms = tim.get((k, float(dt)))
            cells.append(f"{ms*nb/1000:.1f} ({ms:.0f})" if ms else "–")
        L.append(f"| {dt:g} | {nb:,} | {cells[0]} | {cells[1]} | {cells[2]} |")
    L += ["", ME]
    txt = open(DOC, encoding="utf-8").read()
    if MS in txt and ME in txt:
        txt = txt.split(MS)[0] + chr(10).join(L) + txt.split(ME, 1)[1]
        open(DOC, "w", encoding="utf-8").write(txt)


# ── xlsx (수식 오인 수정판) ──
def build_xlsx_safe(data_csv):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    def put(ws, r, c, val):
        cell = ws.cell(r, c)
        cell.value = val
        if isinstance(val, str) and val[:1] in "=+-@":
            cell.data_type = "s"          # 수식 오인 방지 (§10 검증)
        return cell

    def load_sheet(ws, path):
        if not os.path.exists(path):
            return
        rows = list(csv.reader(open(path, encoding="utf-8-sig")))
        for r, row in enumerate(rows, 1):
            for c, v in enumerate(row, 1):
                try:
                    fv = float(v)
                    iv = int(fv)
                    put(ws, r, c, iv if iv == fv and "." not in v and "e" not in v.lower() else fv)
                except ValueError:
                    put(ws, r, c, v)
        for r, row in enumerate(rows, 1):
            first = row[0] if row else ""
            if r == 1 or first.startswith("===") or first in ("bearing", "index"):
                for c in range(1, len(row) + 1):
                    ws.cell(r, c).font = Font(bold=True)
                    if not first.startswith("==="):
                        ws.cell(r, c).fill = PatternFill("solid", fgColor="D9E1F2")
        for c in range(1, max((len(r) for r in rows), default=1) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.freeze_panes = "A2"

    base = os.path.splitext(data_csv)[0]
    wb = Workbook()
    ws1 = wb.active; ws1.title = "데이터"
    load_sheet(ws1, data_csv)
    ws2 = wb.create_sheet("요약")
    load_sheet(ws2, base + "_summary.csv")
    wb.save(base + ".xlsx")


# ── 메인 ──
def main():
    global mf, Design, AnalysisType
    import masta_fatigue as mf          # mastapy.init 포함 (지연 임포트)
    from mastapy.system_model import Design
    from mastapy.system_model.analyses_and_results.static_loads import AnalysisType
    pred = load_pred()
    ref = load_ref()
    data = mf.parse_dlc(mf.DLC_FILE)
    sf, h30 = mf.read_scale_factor(mf.FATIGUE_HRS, DLC_NAME)
    print(f"[참값] UW ISO16281m 표본손상 = {ref:.6e}")

    design = Design.load(MODEL)
    asm = design.all_parts_of_type_root_assembly()[0]
    dp = asm.design_properties
    pl = list(asm.all_parts_of_type_point_load())[0]
    ipl = next(p for p in asm.all_parts_of_type_power_load() if "input" in str(p).lower())
    bearings = list(asm.all_parts_of_type_bearing())
    shaft = list(asm.all_parts_of_type_shaft())[0]
    lc0 = next(c for c in dp.static_loads if c.name == "Load Case 1")
    ds = lc0.design_state_load_case_group
    print("[모델] 로드 완료 · 배치 N =", NBATCH)

    # 웜업 + 자체검증 (최초 1빈: 단일 vs 배치 대조)
    reps0 = bin_reps(data, DTS[0], KS[0])
    mf.set_loads(lc0, pl, ipl, reps0[0][4])
    sd = lc0.analysis_of(AnalysisType.SYSTEM_DEFLECTION)
    sd.perform_analysis()
    l10_single = (sd.results_for(bearings[0]).component_detailed_analysis
                  .iso2812007.basic_rating_life_cycles)
    lc_t = lc0.duplicate(ds, "ks_selfcheck")
    mf.set_loads(lc_t, pl, ipl, reps0[0][4])
    duty_t = dp.add_duty_cycle("ks_selfcheck_dc")
    duty_t.add_static_load(lc_t)
    csd_t = duty_t.analysis_of(AnalysisType.SYSTEM_DEFLECTION)
    csd_t.perform_analysis()
    l10_batch = (list(list(csd_t.results_for(bearings[0]))[0]
                 .component_analysis_cases)[0]
                 .component_detailed_analysis.iso2812007.basic_rating_life_cycles)
    rel = abs(l10_batch / l10_single - 1)
    print(f"[자체검증] 단일 vs 배치 L10 상대오차 = {rel:.2e}  "
          f"{'✅' if rel <= 1e-6 else '❌ 중단'}")
    try:
        lc_t.delete(); duty_t.delete()
    except Exception:
        pass
    if rel > 1e-6:
        return

    # 260722 순서 변경: 저비용 dt 우선(dt-외측), k 교차 — 완료 조합은 요약 존재로 스킵
    COMBOS = [(k, dt) for dt in DTS for k in KS]
    total = len(COMBOS)
    done_n = 0
    for k, dt in COMBOS:
            done_n += 1
            out_csv, sum_csv = paths(k, dt)
            if os.path.exists(sum_csv):
                e = eps_of(sum_csv, ref)
                DONE[(k, dt)] = (e, float("nan"))
                update_doc(f"진행 {done_n}/{total} (기존 재사용)")
                continue
            reps = bin_reps(data, dt, k)
            nb = len(reps)
            f = open(out_csv, "w", newline="", encoding="utf-8-sig")
            w = csv.writer(f)
            w.writerow(mf.DATA_HEADER)
            t_anal = 0.0
            n_done = 0
            for b0 in range(0, nb, NBATCH):
                chunk = reps[b0:b0 + NBATCH]
                cases = []
                for (bi, t_s, rpm, rev, rec) in chunk:
                    lc = lc0.duplicate(ds, f"ks{ktag(k)}_{'%g' % dt}_{bi}")
                    mf.set_loads(lc, pl, ipl, rec)
                    cases.append(lc)
                duty = dp.add_duty_cycle(f"ksdc_{ktag(k)}_{'%g' % dt}_{b0}")
                for lc in cases:
                    duty.add_static_load(lc)
                t0 = time.perf_counter()
                csd = duty.analysis_of(AnalysisType.SYSTEM_DEFLECTION)
                csd.perform_analysis()
                t_anal += time.perf_counter() - t0
                # 케이스별 결과 (베어링 2개)
                subs = {mf.bname(bb): list(list(csd.results_for(bb))[0]
                                           .component_analysis_cases)
                        for bb in bearings}
                for j, (bi, t_s, rpm, rev, rec) in enumerate(chunk):
                    loads = {"force_x_N": -rec["Fz"] * 1e3, "force_y_N": rec["Fy"] * 1e3,
                             "axial_load_N": rec["Fx"] * 1e3, "moment_x_Nm": -rec["Mz"] * 1e3,
                             "moment_y_Nm": rec["My"] * 1e3, "Moment_z_Nm": rec["Mx"] * 1e3}
                    lv = [loads[c] for c in mf.LOAD_COLS]
                    for bb in bearings:
                        d = mf.g(subs[mf.bname(bb)][j], "component_detailed_analysis")
                        sin = mf.fnum(mf.g(d, "maximum_normal_stress_inner"))
                        sout = mf.fnum(mf.g(d, "maximum_normal_stress_outer"))
                        dm = [mf.damage(rev, mf.g(d, pth)) for _, pth, _, _ in mf.DAMAGE_DEFS]
                        w.writerow([bi, t_s, rpm, rev] + lv + [mf.bname(bb),
                                   mf.num(sin / 1e6 if sin is not None else None),
                                   mf.num(sout / 1e6 if sout is not None else None),
                                   mf.num(mf.g(d, "iso762006.safety_factor")),
                                   mf.num(mf.g(d, "iso2812007.basic_rating_life_cycles")),
                                   mf.num(mf.g(d, "iso2812007.modified_rating_life_cycles")),
                                   mf.num(mf.g(d, "isots162812008.basic_reference_rating_life_cycles")),
                                   mf.num(mf.g(d, "isots162812008.modified_reference_rating_life_cycles"))]
                                   + [mf.num(x) for x in dm] + ["", ""])
                    n_done += 1
                for lc in cases:
                    try:
                        lc.delete()
                    except Exception:
                        pass
                try:
                    duty.delete()
                except Exception:
                    pass
                if psutil.virtual_memory().percent > MEM_LIMIT:
                    print(f"  [메모리 {MEM_LIMIT}% 초과] k={k} dt={dt} 배치 중단")
                    break
            f.close()
            mf.write_summary(out_csv, sum_csv, sf, h30, nb)
            try:
                build_xlsx_safe(out_csv)
            except Exception as ex:
                print("  [warn] xlsx 실패:", ex)
            e = eps_of(sum_csv, ref)
            ms = t_anal / max(n_done, 1) * 1000
            DONE[(k, dt)] = (e, ms)
            # timing CSV 갱신 (무상태 update_doc 이 읽는 원천)
            try:
                trows = []
                if os.path.exists(TIMING_CSV):
                    trows = [r for r in csv.DictReader(open(TIMING_CSV, encoding="utf-8-sig"))
                             if not (float(r["k"]) == k and float(r["dt_s"]) == dt)]
                with open(TIMING_CSV, "w", newline="", encoding="utf-8-sig") as tf:
                    tw = csv.writer(tf)
                    tw.writerow(["k", "dt_s", "eps_UW_pct", "ms_per_case"])
                    for r in trows:
                        tw.writerow([r["k"], r["dt_s"], r["eps_UW_pct"], r["ms_per_case"]])
                    tw.writerow([k, dt, f"{e:.4f}" if e is not None else "", f"{ms:.1f}"])
            except Exception as ex:
                print("  [warn] timing 기록 실패:", ex)
            update_doc(f"진행 {done_n}/{total}")
            print(f"[k={k} dt={dt:g}] {nb}빈 해석 {t_anal:.1f}s ({ms:.0f} ms/케이스) "
                  f"ε={e:+.2f}% (예측 {pred.get((k, float(dt)), float('nan')):+.2f}%)")

    update_doc("전체 완료")
    print("\n2단계 스윕 완료")


if __name__ == "__main__":
    if "--doc-only" in sys.argv:
        update_doc("문서 재생성 (--doc-only)")
        print("§9-7 재생성 완료 (MASTA 미사용)")
    else:
        main()
