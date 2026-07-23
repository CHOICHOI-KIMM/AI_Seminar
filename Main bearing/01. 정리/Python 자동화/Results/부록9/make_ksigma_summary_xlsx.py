"""
부록9 총괄 손상정리 엑셀 (부록 8 양식 계승)
===========================================
시트1 총괄      : k{0(빈평균), 0.15, 0.2, 0.25} × dt{10,4,2,1,0.6} + 참값(dt0.1) 손상 168행 + 편향ε
시트2 주지표요약: (k, dt)별 수명·30년손상·a_ISO 평균/σ·편향·**케이스당 해석시간** + 산출절차
데이터: Results/부록8/T50_bin_*  +  Results/부록9/T50_ks*  (MASTA 불필요)
"""
import csv
import glob
import os
import statistics as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
A8 = os.path.join(ROOT, "Results", "부록8")
OUT = os.path.join(HERE, "부록9_총괄_손상정리.xlsx")
DLC = "DLC1.2-c-s1"
DTS = [20, 10, 4, 2, 1, 0.6]
KS = [0.0, 0.15, 0.2, 0.25]          # 0.0 = 빈 산술평균 (부록 8)
REF_DT = 0.1
PRIMARY = ("Main Bearing_UW", "ISO16281", "modified")
TIMING = os.path.join(HERE, "ksigma_timing.csv")

HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
REF_FILL = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
TH = Side(style="thin", color="BFBFBF")
BD = Border(left=TH, right=TH, top=TH, bottom=TH)
BLUE, RED = Font(color="1F60C4"), Font(color="C00000")


def summary_path(k, dt):
    if dt == REF_DT or k == 0.0:
        return os.path.join(A8, f"T50_bin_{DLC}_dt_{'%g' % dt}_summary.csv")
    return os.path.join(HERE, f"T50_ks{'%g' % k}_{DLC}_dt_{'%g' % dt}_summary.csv")


def data_path(k, dt):
    return summary_path(k, dt).replace("_summary.csv", ".csv")


def load_a(path):
    rows, insec, hdr = [], False, None
    for r in csv.reader(open(path, encoding="utf-8-sig")):
        if not r:
            continue
        c0 = r[0].lstrip("﻿")
        if c0.startswith("=== (a)"):
            insec = True; continue
        if c0.startswith("=== ("):
            insec = False; continue
        if not insec:
            continue
        if hdr is None and c0 == "bearing":
            hdr = r; continue
        if hdr and c0.startswith("Main"):
            rows.append(dict(zip(hdr, r)))
    return rows


def aiso_stats(path):
    """데이터 CSV → {bearing: (평균, σ)} (ISO16281 a_ISO = L10mr/L10r)."""
    acc = {}
    for r in csv.DictReader(open(path, encoding="utf-8-sig")):
        try:
            l10r = float(r["L10r_rev"])
            if l10r > 0:
                acc.setdefault(r["bearing"], []).append(float(r["L10mr_rev"]) / l10r)
        except (KeyError, ValueError):
            pass
    return {b: (st.mean(v), st.pstdev(v) if len(v) > 1 else 0.0) for b, v in acc.items()}


def load_timing():
    t = {}
    if os.path.exists(TIMING):
        for r in csv.DictReader(open(TIMING, encoding="utf-8-sig")):
            t[(float(r["k"]), float(r["dt_s"]))] = float(r["ms_per_case"])
    return t


def n_bins(dt):
    return 6001 if dt == REF_DT else int(600 / dt)


def style_eps(cell, v):
    if v is None:
        cell.value = "–"; return
    cell.value = v / 100.0
    cell.number_format = "+0.00%;-0.00%"
    cell.font = BLUE if v < 0 else RED


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def klabel(k, dt):
    if dt == REF_DT:
        return "참값(dt0.1)"
    return "빈평균(k=0)" if k == 0.0 else f"평균+{('%g' % k)}σ"


def main():
    data = {}          # {(k,dt): {(b,std,basis): row}}
    combos = [(0.0, REF_DT)] + [(k, dt) for k in KS for dt in DTS]
    for k, dt in combos:
        p = summary_path(k, dt)
        if not os.path.exists(p):
            print("[누락]", p); continue
        data[(k, dt)] = {(r["bearing"], r["standard"], r["life_basis"]): r
                         for r in load_a(p)}
    ref = data[(0.0, REF_DT)]
    refUW = float(ref[PRIMARY]["sample_damage"])
    refDW = float(ref[("Main Bearing_DW", "ISO16281", "modified")]["sample_damage"])
    _E = 9.0 / 8.0
    _rlu = float(ref[PRIMARY]["equiv_life_years"])
    _rld = float(ref[("Main Bearing_DW", "ISO16281", "modified")]["equiv_life_years"])
    refLS = (_rlu ** -_E + _rld ** -_E) ** (-1 / _E)
    timing = load_timing()

    wb = Workbook(); wb.remove(wb.active)

    # ── 시트1 총괄 ──
    ws = wb.create_sheet("총괄")
    cols = ["k", "dt_s", "n_bins", "bearing", "standard", "life_basis",
            "sample_damage", "scale_factor", "damage_30yr", "SF_fatigue",
            "equiv_life_years", "편향_ε_%(vs dt0.1)", "note"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c); cell.font = BOLD; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = BD
    keys = [(b, s, lb) for b in ("Main Bearing_UW", "Main Bearing_DW")
            for s in ("ISO281", "ISO16281") for lb in ("basic", "modified")]
    r = 2
    for k, dt in combos:
        if (k, dt) not in data:
            continue
        for key in keys:
            row = data[(k, dt)].get(key)
            if not row:
                continue
            sd = float(row["sample_damage"])
            rv = ref.get(key)
            ev = (sd / float(rv["sample_damage"]) - 1) * 100 if rv else None
            vals = [klabel(k, dt), dt, n_bins(dt), key[0], key[1], key[2],
                    sd, int(float(row["scale_factor"])), float(row["damage_30yr"]),
                    float(row["SF_fatigue"]), float(row["equiv_life_years"]),
                    None, row.get("note", "")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c)
                if c == 12:
                    style_eps(cell, ev)
                else:
                    cell.value = v
                cell.border = BD
            ws.cell(r, 7).number_format = "0.000E+00"
            ws.cell(r, 9).number_format = "0.0000"
            ws.cell(r, 10).number_format = "0.000"
            ws.cell(r, 11).number_format = "#,##0.0"
            if dt == REF_DT:
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = REF_FILL
            r += 1
    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r-1}"
    autosize(ws, [13, 7, 8, 17, 10, 11, 13, 11, 12, 11, 15, 16, 16])
    n1 = r - 2

    # ── 시트2 주지표요약 ──
    ws2 = wb.create_sheet("주지표요약")
    ws2.append([f"부록 9 · ISO16281 수정수명 (윤활유 50°C, {DLC}) — 대표값 방법별"])
    ws2.cell(1, 1).font = Font(bold=True, size=12)
    hdr = ["방법", "k", "dt (s)", "빈수", "UW수명(yr)", "DW수명(yr)", "L10,sys(yr)",
           "UW 30년손상", "DW 30년손상", "UW a_ISO 평균", "UW a_ISO σ",
           "DW a_ISO 평균", "DW a_ISO σ",
           "편향 ε (UW)", "편향 ε (DW)", "편향 ε (Sys)", "케이스당 해석 [ms]", "총 해석시간 [s]"]
    ws2.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws2.cell(2, c); cell.font = BOLD; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = BD
    K = lambda b: (b, "ISO16281", "modified")
    r = 3
    for k, dt in combos:
        if (k, dt) not in data:
            continue
        d = data[(k, dt)]
        ai_all = aiso_stats(data_path(k, dt))
        ai = ai_all.get("Main Bearing_UW", (None, None))
        aid = ai_all.get("Main Bearing_DW", (None, None))
        uw, dw = d.get(K("Main Bearing_UW")), d.get(K("Main Bearing_DW"))
        ev = (float(uw["sample_damage"]) / refUW - 1) * 100 if uw else None
        ms = timing.get((k, dt))
        E_W = 9.0 / 8.0
        lu = float(uw["equiv_life_years"]) if uw else None
        ld = float(dw["equiv_life_years"]) if dw else None
        lsys = (lu ** -E_W + ld ** -E_W) ** (-1 / E_W) if (lu and ld) else None
        evD = (float(dw["sample_damage"]) / refDW - 1) * 100 if dw else None
        evS = (refLS / lsys - 1) * 100 if lsys else None      # 손상 관점 (참값수명/수명−1)
        vals = [klabel(k, dt), k if dt != REF_DT else "–", dt, n_bins(dt),
                lu, ld, lsys,
                float(uw["damage_30yr"]) if uw else None,
                float(dw["damage_30yr"]) if dw else None,
                ai[0], ai[1], aid[0], aid[1],
                None, None, None, ms if ms is not None else "–",
                (ms * n_bins(dt) / 1000.0) if ms is not None else "–"]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c)
            if c in (14, 15, 16):
                e = {14: ev, 15: evD, 16: evS}[c]
                style_eps(cell, e if dt != REF_DT else 0.0)
            else:
                cell.value = v
            cell.border = BD
        for c, fmt in ((5, "#,##0.0"), (6, "#,##0.0"), (7, "#,##0.0"), (8, "0.0000"),
                       (9, "0.0000"), (10, "0.0000"), (11, "0.0000"),
                       (12, "0.0000"), (13, "0.0000"), (17, "0"), (18, "0.0")):
            ws2.cell(r, c).number_format = fmt
        if dt == REF_DT:
            for c in range(1, len(hdr) + 1):
                ws2.cell(r, c).fill = REF_FILL
        r += 1
    ws2.freeze_panes = "D3"
    autosize(ws2, [13, 6, 8, 8, 13, 13, 13, 12, 12, 13, 12, 13, 12, 12, 12, 12, 15, 13])

    # 하단 산출 절차 (부록 8 양식 계승)
    mono = Font(name="Consolas", size=10)
    note_it = Font(italic=True, color="808080", size=9)
    r += 2
    ws2.cell(r, 1, "■ 산출 절차 (부록 8 총괄 양식 계승)")
    ws2.cell(r, 1).font = Font(bold=True, size=11)
    ws2.cell(r, 1).fill = PatternFill("solid", fgColor="EAECEE")
    proc = [
        ("[1] 대표하중: 빈 내 5분력 각각 rep = mu + sign(mu)·k·sigma (sigma=모표준편차 ÷n). "
         "M_Z·rpm 은 빈 산술평균. k=0 이 부록 8 빈평균과 동일", "m"),
        ("[2] 해석: MASTA System Deflection, 배치 N=20 (duplicate → duty cycle → "
         "component_analysis_cases 추출, §10-6.1)", "m"),
        ("[3] 손상: D = Σ N_b / L10mr,  N_b = (rpm_avg/60)·dt (외부 Miner, MASTA Damage% 미사용)", "m"),
        ("[4] a_ISO,i = L10mr_i / L10r_i → 조건별 평균·모표준편차(÷n)", "m"),
        ("[5] 편향 ε = (표본손상 − 참값)/참값,  참값 = dt 0.1 점별(부록 8 T50_bin)", "m"),
        ("[6] 케이스당 해석시간 = 배치 해석 벽시계시간 / 케이스수 (ksigma_timing.csv)", "m"),
        ("[7] 시스템 수명 L10,sys = [L_UW^(-9/8) + L_DW^(-9/8)]^(-8/9)  "
         "(Lundberg-Palmgren 직렬 와이블, e=9/8, 독립 가정)", "m"),
        ("[8] 편향 ε: UW·DW = 표본손상 비 − 1,  Sys = 참값L10,sys/L10,sys − 1  "
         "(모두 손상 관점, +=보수측)", "m"),
        ("[9] 총 해석시간 [s] = 케이스당 해석 [ms] × 빈수 / 1000  (배치 준비시간 제외)", "m"),
        ("※ 합격 기준: 0 ≤ ε ≤ +3% (보수측 + 3% 이내). k=0.2 가 전 dt 유일 합격 (§9-8)", "i"),
    ]
    for txt, sty in proc:
        r += 1
        c = ws2.cell(r, 1, txt)
        c.font = mono if sty == "m" else note_it

    wb.save(OUT)
    print(f"[저장] {OUT}")
    print(f"  시트1 총괄: {n1}행 · 시트2 주지표요약: {len(combos)}조합")


if __name__ == "__main__":
    main()
