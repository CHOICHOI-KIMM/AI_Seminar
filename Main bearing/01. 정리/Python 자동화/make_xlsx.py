"""fatigue 데이터 CSV + 요약 CSV → 단일 .xlsx(데이터·요약 2시트). MASTA 불필요.
사용: python make_xlsx.py <data_csv>   (없으면 아래 DEFAULT)
"""
import csv
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "fatigue_test10d_DLC1.2-c-s1.csv")


def num_or_str(s):
    if s == "" or s is None:
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def load_sheet(ws, csv_path, header_fill="D9E1F2"):
    if not os.path.exists(csv_path):
        ws["A1"] = f"(파일 없음: {os.path.basename(csv_path)})"
        return
    with open(csv_path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, num_or_str(val))
    # 헤더/섹션 강조
    for r, row in enumerate(rows, 1):
        first = (row[0] if row else "")
        is_sec = first.startswith("===")
        is_hdr = (r == 1) or is_sec or (row and row[0] in ("bearing", "index",
                  "min_shaft_DIN743_SF_inf"))
        if is_hdr:
            for c in range(1, len(row) + 1):
                cell = ws.cell(r, c)
                cell.font = Font(bold=True)
                if not is_sec:
                    cell.fill = PatternFill("solid", fgColor=header_fill)
    # 열 너비
    if rows:
        for c in range(1, max(len(r) for r in rows) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"


def build_xlsx(data_csv):
    """데이터 CSV + (자동 유추)요약 CSV → 단일 xlsx. 생성경로 반환."""
    base = os.path.splitext(data_csv)[0]
    summary_csv = base + "_summary.csv"
    out_xlsx = base + ".xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "데이터"
    load_sheet(ws1, data_csv)
    ws2 = wb.create_sheet("요약")
    load_sheet(ws2, summary_csv)
    wb.save(out_xlsx)
    print("엑셀 생성:", out_xlsx)
    return out_xlsx


def main():
    build_xlsx(sys.argv[1] if len(sys.argv) > 1 else DEFAULT)


if __name__ == "__main__":
    main()
